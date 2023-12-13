from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import *
from .forms import ContactForm, CommentForm
from django.core.mail import BadHeaderError, send_mail
from django.http import JsonResponse
from django.core import serializers
from django.contrib.auth.views import LogoutView
from django.contrib.auth import logout
from django.contrib.auth.models import User
from openpyxl import load_workbook
from datetime import datetime, timedelta
from django.contrib import messages
from django.core.paginator import Paginator, Page


allowedEmails=["harsh247gupta@gmail.com", "harsh90731@gmail.com", "rajumeshram767@gmail.com", "hariomk628@gmail.com"]

# def index(request):
#     return HttpResponse("Hello, world")

#code to import excel entries into azad_boarders model
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            roll_no, name, email, number = row
            azad_boarders.objects.create(name=name, roll_no=roll_no, emails=email, contact=number)

        return render(request, 'index.html')
def addBoarders(request):
    if request.user.is_authenticated:
        email=request.user.email
        if email in allowedEmails:
            return render(request, 'addBoarders.html')
    messages.info(request, 'Please login with valid ID to add boarders')
    return redirect("/")
    
     


def login(request):
    return render(request, 'login.html')

def index(request):
    if request.user.is_authenticated:
        email=request.user.email
        boarders=azad_boarders.objects.all()
        for boarder in boarders:
            if boarder.emails==email:
                name=boarder.name
                params={"name":name}
                return render(request, 'index.html', params)
        logout(request)
        message="Please login with valid EmailID"
        params={"message":message}
        return render(request, 'index.html', params)
    
    return render(request, 'index.html')

def complain(request):
    if request.user.is_authenticated:
        return render(request, 'complain.html')


def submit_form(request):
    if request.method == 'POST' and request.user.is_authenticated :
        category = request.POST.get('category')
        room_no = request.POST.get('room_no')
        complain = request.POST.get('complain')
        contact_no  = request.POST.get('contact_no')
        image = request.FILES.get('image')
        boarder = azad_boarders.objects.get(emails = request.user.email)
        name=boarder.name
        email= boarder.emails
        roll_no = boarder.roll_no
        now=datetime.now()
        t_string = now.strftime("%d/%m/%Y %H:%M %p")
        created_at=t_string
        register = complaints.objects.create(name=name, roll_no=roll_no, email=email, category=category, contact_no=contact_no, complain=complain, status="pending", room_no=room_no, created_at=created_at, image=image, review="None")
        # Return a response (you can render a template or return a JSON response)
        # message="Complain submitted successfully!"
        # params={"message":message,"name":name}
        messages.info(request, 'Complain submitted successfully')
        return redirect("/")
    else:
        # Handle GET requests or other methods if necessary
        return HttpResponse('Invalid request method')

def showComplaints(request):
    
    if request.user.is_authenticated:
        email=request.user.email
        if email in allowedEmails:
            pending_complaints = complaints.objects.filter(status="pending")
            pending_paginator = Paginator(pending_complaints, 10)
            current_page_pending = pending_paginator.page(request.GET.get('pending_page', 1))

            completed_complaints = complaints.objects.filter(status="Completed").order_by("-id")
            ongoing_complaints = complaints.objects.filter(status="Ongoing")
            params = {"pending_complaints": current_page_pending, "completed_complaints": completed_complaints, "ongoing_complaints": ongoing_complaints}
            return render(request,'showComplaints.html', params)
    messages.info(request, 'Please login with valid ID to access complaints')
    return redirect("/")
    

def showFullComplain(request, complain_id):
    complain=complaints.objects.get(id=complain_id)
    params = {"complain": complain}
    return render(request,"fullComplain.html", params)


def updateStatus(request):
    if request.method=="POST":
        id = request.POST.get('id')
        manager_review = request.POST.get('manager_review')
        set_status = request.POST.get('set_status')
        print(set_status)
        if manager_review=="":
            manager_review="None"
        complain = complaints.objects.get(id=id)
        if(complain.status=="Ongoing"):
            complain.status="Completed"
        else:
            if(set_status=="ongoing"):
                complain.status="Ongoing"
            else:
                complain.status="Completed"
        # if(complain.status=="pending"):
        #     complain.status="Ongoing"
        # else:
        #     complain.status="Completed"
        complain.manager_review=manager_review
        now=datetime.now()
        t_string = now.strftime("%d/%m/%Y %H:%M %p")
        complain.modified_at=t_string
        complain.save()
    return redirect("/showComplaints")

def complain_status(request):
    complains = complaints.objects.filter(email=request.user.email).order_by("-id")
    print(complains)
    params={"complains":complains}
    return render(request, "complain_status.html", params)

def noticeboard(request):
    noticeboard = Notice.objects.all()
    return render(request, 'noticeboard.html', {'noticeboard': noticeboard})

def notice(request, noticeid):
    notice = Notice.objects.filter(id=noticeid)
    noticeboard = Notice.objects.all()
    return render(request, 'notice.html', {'notice': notice[0], 'noticeboard': noticeboard})


def achievements(request):
    achievements = Achievements.objects.all()
    return render(request, 'achievements.html', {'achievements': achievements})


def achievement(request, achievementid):
    achievement = Achievements.objects.filter(id=achievementid)
    achievements = Achievements.objects.all()
    return render(request, 'achievement.html', {'achievement': achievement[0], "achievements": achievements})

def events(request):
    events = Event.objects.all()
    return render(request, 'events.html', {'events': events})

def khoj(request):
     return render(request, 'khoj.html')

def about(request):
    if (request.method == 'POST'):
        form = ContactForm(request.POST)
        form.save()
        subject = request.POST.get('subject', '')
        message = request.POST.get('msg', '')
        from_email = request.POST.get('email', '')
        if subject and message and from_email:
            try:
                send_mail(subject, message, from_email,
                          ['smarakdev@gmail.com'])
            except BadHeaderError:
                return HttpResponse('Invalid header found.')
            return redirect('/about')
        else:
            form = ContactForm()
            return render(request, 'about.html', {'form': form})
    else:
        form = ContactForm()
        return render(request, 'about.html', {'form': form})


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


def postComment(request):
    # request should be ajax and method should be POST.
    print("hello")
    if is_ajax(request=request) and request.method == "POST":
        # get the form data
        form = CommentForm(request.POST)
        # save the data and after fetch the object in instance
        if form.is_valid():
            instance = form.save()
            # serialize in new friend object in json
            ser_instance = serializers.serialize('json', [instance, ])
            # send to client side.
            print("saved")
            return JsonResponse({"instance": ser_instance}, status=200)
        else:
            # some form errors occured.
            return JsonResponse({"error": form.errors}, status=400)

    # some error occured
    return JsonResponse({"error": ""}, status=400)


def event(request, eventid):
    id = eventid
    # print("message", eventid)
    event = Event.objects.filter(id=eventid)
    itemlist = []
    # print(event)
    # print(event[0].para_set.all())
    commentform = CommentForm()
    comments = event[0].comments.all()

    cover = ""
    for item in event[0].imagemodel_set.all():
        if item.caption == 'cover':
            cover = item
            continue
        itemlist.append(item)
    for item in event[0].para_set.all():
        itemlist.append(item)
    # print(itemlist)
    itemlist.sort(key=lambda x: x.date, reverse=True)
    # print(itemlist)
    # print(cover.caption)
    events = Event.objects.all()
    return render(request, 'event.html', {'events': events, 'cover': cover, 'event': event[0], 'form': commentform, 'comments': comments})

def custom_logout(request):
    logout(request)
    message="Logged out successfully"
    params={"message":message}
    return render(request,"index.html", params)

